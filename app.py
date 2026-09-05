from dotenv import load_dotenv
load_dotenv()
load_dotenv(".env.local")
import os
import csv
import io
import sqlite3
import secrets
import hashlib
import base64
import unicodedata
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path
from urllib.parse import urlparse, unquote

def normalize_headers(row):
    headers = []
    for c in row:
        val = str(c or "").strip()
        val = unicodedata.normalize('NFKD', val).encode('ASCII', 'ignore').decode('utf-8')
        headers.append(val.lower())
    return headers

def find_header_and_data_rows(all_rows):
    if not all_rows:
        return 0, [], []
    candidate_keywords = ["produto", "product", "modelo", "model", "nome", "name", "sku", "codigo", "code", "varejo", "retail", "atacado", "wholesale"]
    header_idx = 0
    for idx, row in enumerate(all_rows[:10]):
        norm = normalize_headers(row)
        if any(kw in norm for kw in candidate_keywords) or any(any(kw in cell for kw in candidate_keywords) for cell in norm):
            header_idx = idx
            break
    headers = normalize_headers(all_rows[header_idx])
    return header_idx, headers, all_rows[header_idx + 1:]



from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory, jsonify
from werkzeug.utils import secure_filename

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from gemini_service import (
    ask_gemini_copilot,
    get_gemini_api_key,
    extract_and_match_chassis,
    analyze_payment_receipt,
    analyze_import_documents,
    PAYMENT_CATEGORIES,
    ACCOUNTS_LIST,
    PAYMENT_METHODS
)
import bling_service


import base64
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:
    psycopg2 = None

DEFAULT_DB_URL = "postgresql://postgres.ztbmnzwrpigcohwobrig:%40Jammajjam24@aws-0-us-west-2.pooler.supabase.com:6543/postgres?sslmode=require"
DATABASE_URL = os.environ.get("DATABASE_URL") or os.environ.get("SUPABASE_DB_URL") or DEFAULT_DB_URL
pg_pool = None

def get_pg_pool():
    global pg_pool
    db_url = os.environ.get("DATABASE_URL") or os.environ.get("SUPABASE_DB_URL") or DEFAULT_DB_URL
    if db_url and psycopg2 and pg_pool is None:
        try:
            pg_pool = psycopg2.pool.ThreadedConnectionPool(1, 10, dsn=db_url)
            print("POSTGRESQL CONNECTION POOL INITIALIZED!")
        except Exception as e:
            print("Failed to initialize PG pool:", e)
            pg_pool = None
    return pg_pool

class PGCursorWrapper:
    def __init__(self, cur, lastrowid=None):
        self.cur = cur
        self.lastrowid = lastrowid
    def fetchone(self):
        return self.cur.fetchone()
    def fetchall(self):
        return self.cur.fetchall()
    def __iter__(self):
        return iter(self.cur)

class PGConnWrapper:
    def __init__(self, conn, pool=None):
        self.conn = conn
        self.pool = pool
    def execute(self, sql, params=()):
        pg_sql = sql.replace("?", "%s")
        pg_sql = pg_sql.replace("active=1", "active=TRUE").replace("active=0", "active=FALSE")
        pg_sql = pg_sql.replace("promo_eligible=1", "promo_eligible=TRUE").replace("promo_eligible=0", "promo_eligible=FALSE")
        pg_sql = pg_sql.replace("COALESCE(st.received_at, substr(st.created_at,1,10))", "COALESCE(st.received_at, st.created_at::date)")
        pg_sql = pg_sql.replace("HAVING available>0 AND oldest_date<=", "HAVING COUNT(st.id)>0 AND MIN(COALESCE(st.received_at, st.created_at::date))<=")
        pg_sql = pg_sql.replace("GROUP_CONCAT(st.chassis, ', ')", "STRING_AGG(st.chassis, ', ')")
        
        is_insert = pg_sql.strip().upper().startswith("INSERT INTO")
        if is_insert and "RETURNING" not in pg_sql.upper():
            pg_sql += " RETURNING id"
        
        cur = self.conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(pg_sql, params)
        
        lastrowid = None
        if is_insert:
            try:
                row = cur.fetchone()
                if row and "id" in row:
                    lastrowid = row["id"]
            except Exception:
                pass
        return PGCursorWrapper(cur, lastrowid)

    def executescript(self, sql):
        cur = self.conn.cursor()
        cur.execute(sql)
        return cur
    def commit(self):
        self.conn.commit()
    def __enter__(self):
        return self
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            self.conn.rollback()
        else:
            self.conn.commit()
        if self.pool:
            self.pool.putconn(self.conn)
        else:
            self.conn.close()

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "m_one.db"
if os.environ.get("VERCEL"):
    UPLOAD_DIR = Path("/tmp/uploads")
else:
    UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or os.environ.get("M_ONE_SECRET") or "maj-m-one-production-fixed-secret-key-2026-v1"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

@app.after_request
def add_no_cache_headers(response):
    if "text/html" in response.headers.get("Content-Type", ""):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "webp", "csv", "xlsx", "xls"}

ROLE_LABELS = {
    "admin": "Diretoria",
    "finance": "Financeiro",
    "stock": "Estoque",
    "sales": "Vendas",
    "support": "Suporte Técnico",
}


DEFAULT_DB_URL = "postgresql://postgres.ztbmnzwrpigcohwobrig:%40Jammajjam24@aws-0-us-west-2.pooler.supabase.com:6543/postgres"

def connect_pg(db_url: str):
    if not db_url or not psycopg2:
        return None
    try:
        parsed = urlparse(db_url)
        user = unquote(parsed.username) if parsed.username else None
        pwd = unquote(parsed.password) if parsed.password else None
        dbname = parsed.path.lstrip("/") if parsed.path else "postgres"
        host = parsed.hostname
        port = parsed.port or 5432
        return psycopg2.connect(
            host=host,
            port=port,
            user=user,
            password=pwd,
            dbname=dbname,
            sslmode="require",
            connect_timeout=10
        )
    except Exception as e:
        print("PG connect_pg failed:", e)
        # Fallback to raw string
        return psycopg2.connect(db_url, sslmode="require", connect_timeout=10)

def db():
    db_url = os.environ.get("DATABASE_URL") or os.environ.get("SUPABASE_DB_URL") or DEFAULT_DB_URL
    if db_url and psycopg2:
        try:
            conn = connect_pg(db_url)
            if conn:
                return PGConnWrapper(conn)
        except Exception as e:
            print("Direct PG connection failed:", e)

    if os.environ.get("VERCEL"):
        tmp_db_path = Path("/tmp/m_one.db")
        conn = sqlite3.connect(tmp_db_path)
    else:
        conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn




def hash_password(password: str) -> str:
    salt = "m-one-v1"
    return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()


def init_db():
    if os.environ.get("DATABASE_URL") or os.environ.get("SUPABASE_DB_URL") or DEFAULT_DB_URL:
        # PostgreSQL / Supabase tables are initialized via Supabase SQL Editor
        return
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'sales',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sku TEXT UNIQUE,
                category TEXT,
                unit_cost REAL NOT NULL DEFAULT 0,
                retail_price REAL NOT NULL DEFAULT 0,
                wholesale_price REAL NOT NULL DEFAULT 0,
                promo_eligible INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reference TEXT NOT NULL UNIQUE,
                invoice_no TEXT,
                bl_no TEXT,
                arrival_date TEXT,
                usd_rate REAL NOT NULL DEFAULT 0,
                invoice_amount_usd REAL NOT NULL DEFAULT 0,
                nf_entry TEXT,
                invoice_file TEXT,
                bl_file TEXT,
                nf_entry_file TEXT,
                chassis_file TEXT,
                notes TEXT,
                status TEXT NOT NULL DEFAULT 'draft',
                created_by INTEGER,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(created_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS import_costs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                import_id INTEGER NOT NULL,
                cost_type TEXT NOT NULL,
                description TEXT,
                amount REAL NOT NULL,
                currency TEXT NOT NULL DEFAULT 'BRL',
                usd_rate REAL NOT NULL DEFAULT 0,
                paid_at TEXT,
                receipt_file TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(import_id) REFERENCES imports(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS stock_units (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chassis TEXT NOT NULL UNIQUE,
                motor_no TEXT,
                product_id INTEGER NOT NULL,
                color TEXT,
                import_id INTEGER,
                status TEXT NOT NULL DEFAULT 'available',
                location TEXT NOT NULL DEFAULT 'Depósito',
                received_at TEXT,
                sold_at TEXT,
                sale_id INTEGER,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(product_id) REFERENCES products(id),
                FOREIGN KEY(import_id) REFERENCES imports(id)
            );

            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_number TEXT NOT NULL,
                invoice_number TEXT NOT NULL,
                channel TEXT NOT NULL,
                customer TEXT,
                sold_at TEXT NOT NULL,
                total_value REAL NOT NULL DEFAULT 0,
                notes TEXT,
                created_by INTEGER,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(created_by) REFERENCES users(id)
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_sales_invoice_unique ON sales(invoice_number);

            CREATE TABLE IF NOT EXISTS sale_units (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL,
                stock_unit_id INTEGER NOT NULL UNIQUE,
                product_id INTEGER NOT NULL,
                unit_value REAL NOT NULL DEFAULT 0,
                FOREIGN KEY(sale_id) REFERENCES sales(id) ON DELETE CASCADE,
                FOREIGN KEY(stock_unit_id) REFERENCES stock_units(id),
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            CREATE TABLE IF NOT EXISTS sale_receipts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL,
                method TEXT NOT NULL,
                account TEXT,
                amount REAL NOT NULL,
                received_at TEXT NOT NULL,
                receipt_file TEXT,
                FOREIGN KEY(sale_id) REFERENCES sales(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                paid_at TEXT NOT NULL,
                description TEXT NOT NULL,
                category TEXT,
                amount REAL NOT NULL,
                account TEXT,
                receipt_file TEXT,
                import_id INTEGER,
                visibility TEXT NOT NULL DEFAULT 'finance',
                created_by INTEGER,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(import_id) REFERENCES imports(id),
                FOREIGN KEY(created_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT NOT NULL,
                detail TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(user_id) REFERENCES users(id)
            );
            """
        )
        # Default users only on first run.
        if conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"] == 0:
            defaults = [
                ("Jean", "jean", hash_password("MOne2026!"), "admin"),
                ("Geysa", "geysa", hash_password("MOne2026!"), "admin"),
                ("Marisa", "marisa", hash_password("MOne2026!"), "finance"),
                ("Jhon", "jhon", hash_password("MOne2026!"), "stock"),
                ("Luísa", "luisa", hash_password("MOne2026!"), "sales"),
                ("Léo", "leo", hash_password("MOne2026!"), "sales"),
                ("Gabriel", "gabriel", hash_password("MOne2026!"), "sales"),
            ]
            conn.executemany("INSERT INTO users(name,username,password_hash,role) VALUES(?,?,?,?)", defaults)
        conn.commit()
    ensure_sales_columns()
    ensure_product_columns()
    ensure_payments_columns()

def ensure_sales_columns():
    try:
        with db() as conn:
            for col in ["danfe_file", "delivery_term_files", "ai_chassis_verified", "ai_extracted_chassis", "vehicle_model", "chassis_photo_file", "warranty_term_file", "signed_stub_file"]:
                try:
                    conn.execute(f"ALTER TABLE sales ADD COLUMN {col} TEXT")
                except Exception:
                    pass
            conn.commit()
    except Exception:
        pass

def ensure_product_columns():
    try:
        with db() as conn:
            for col, col_type in [
                ("bling_stock", "INTEGER DEFAULT 0"),
                ("bling_updated_at", "TEXT"),
                ("fob_price_usd", "REAL DEFAULT 0"),
                ("aliquota_rate", "REAL DEFAULT 0"),
            ]:
                try:
                    conn.execute(f"ALTER TABLE products ADD COLUMN {col} {col_type}")
                except Exception:
                    pass
            conn.commit()
    except Exception:
        pass

def ensure_payments_columns():
    try:
        with db() as conn:
            for col, col_type in [
                ("payment_method", "TEXT"),
                ("card_last4", "TEXT"),
                ("supplier", "TEXT"),
                ("document_no", "TEXT"),
                ("ai_verified", "INTEGER DEFAULT 0")
            ]:
                try:
                    conn.execute(f"ALTER TABLE payments ADD COLUMN {col} {col_type}")
                except Exception:
                    pass
            conn.commit()
    except Exception:
        pass
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS



def save_base64_upload(base64_str, prefix="photo"):
    if not base64_str or "," not in base64_str:
        return None
    try:
        header, data = base64_str.split(",", 1)
        raw_bytes = base64.b64decode(data)
        ext = "png" if "png" in header else "jpg"
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = f"{prefix}_{stamp}.{ext}"
        (UPLOAD_DIR / filename).write_bytes(raw_bytes)
        return filename
    except Exception:
        return None

def save_upload(file_storage, prefix="file"):
    if not file_storage or not file_storage.filename:
        return None
    if not allowed_file(file_storage.filename):
        raise ValueError("Tipo de arquivo não permitido")
    original = secure_filename(file_storage.filename)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    filename = f"{prefix}_{stamp}_{original}"
    file_storage.save(UPLOAD_DIR / filename)
    return filename


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    with db() as conn:
        return conn.execute("SELECT * FROM users WHERE id=? AND active=1", (uid,)).fetchone()


def login_required(fn):
    @wraps(fn)
    def inner(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return inner


def roles_required(*roles):
    def decorator(fn):
        @wraps(fn)
        def inner(*args, **kwargs):
            u = current_user()
            if not u or u["role"] not in roles:
                flash("Você não tem permissão para acessar esta área.", "danger")
                return redirect(url_for("dashboard"))
            return fn(*args, **kwargs)
        return inner
    return decorator


def audit(action, detail=""):
    with db() as conn:
        conn.execute("INSERT INTO audit_log(user_id,action,detail) VALUES(?,?,?)", (session.get("user_id"), action, detail))
        conn.commit()


def money(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def money_usd(v):
    try:
        return f"$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "$ 0,00"


def aliquota(v):
    try:
        if not v or float(v) == 0:
            return "—"
        return f"R$ {float(v):,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


app.jinja_env.filters["money"] = money
app.jinja_env.filters["money_usd"] = money_usd
app.jinja_env.filters["aliquota"] = aliquota


@app.context_processor
def inject_globals():
    return {"me": current_user(), "role_labels": ROLE_LABELS}


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        try:
            with db() as conn:
                user = conn.execute("SELECT * FROM users WHERE lower(username)=lower(?) AND (active=1 OR active IS TRUE)", (username,)).fetchone()
            if user and user["password_hash"] == hash_password(password):
                session.permanent = True
                session["user_id"] = user["id"]
                flash(f"Bem-vindo, {user['name']}.", "success")
                return redirect(url_for("dashboard"))
            flash("Usuário ou senha inválidos.", "danger")
        except Exception as e:
            flash(f"Erro de conexão com o banco de dados: {str(e)}", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    today = date.today()
    month_start = today.replace(day=1).isoformat()
    today_iso = today.isoformat()
    ago90 = (today - timedelta(days=90)).isoformat()
    ago30 = (today - timedelta(days=30)).isoformat()

    with db() as conn:
        st_row = conn.execute("SELECT COALESCE(SUM(total_value),0) v, COUNT(*) c FROM sales WHERE sold_at=?", (today_iso,)).fetchone()
        sales_today = st_row["v"]
        sales_today_count = st_row["c"]
        sales_month = conn.execute("SELECT COALESCE(SUM(total_value),0) v FROM sales WHERE sold_at>=?", (month_start,)).fetchone()["v"]
        payments_month = conn.execute("SELECT COALESCE(SUM(amount),0) v FROM payments WHERE paid_at>=?", (month_start,)).fetchone()["v"]
        stock_available = conn.execute("SELECT COUNT(*) c FROM stock_units WHERE status='available'").fetchone()["c"]
        top_products = conn.execute(
            """
            SELECT p.id,p.name,COUNT(su.id) units,COALESCE(SUM(su.unit_value),0) revenue
            FROM sale_units su JOIN sales s ON s.id=su.sale_id JOIN products p ON p.id=su.product_id
            WHERE s.sold_at>=?
            GROUP BY p.id,p.name ORDER BY units DESC,revenue DESC LIMIT 6
            """, (month_start,)
        ).fetchall()
        opportunities = conn.execute(
            """
            SELECT p.id,p.name,p.unit_cost,p.wholesale_price,p.retail_price,
                   COUNT(st.id) available,
                   MIN(COALESCE(st.received_at, substr(st.created_at,1,10))) oldest_date,
                   COALESCE((SELECT COUNT(*) FROM sale_units su2 JOIN sales s2 ON s2.id=su2.sale_id
                             WHERE su2.product_id=p.id AND s2.sold_at>=?),0) sold_30
            FROM products p JOIN stock_units st ON st.product_id=p.id AND st.status='available'
            WHERE p.promo_eligible=1
            GROUP BY p.id
            HAVING available>0 AND oldest_date<=?
            ORDER BY sold_30 ASC, oldest_date ASC, available DESC LIMIT 6
            """, (ago30, ago90)
        ).fetchall()
        chassis_alerts = conn.execute(
            """
            SELECT COUNT(*) c FROM stock_units st
            LEFT JOIN imports i ON i.id=st.import_id
            WHERE st.status='available' AND (i.id IS NULL OR i.status!='released')
            """
        ).fetchone()["c"]

    opp = []
    for r in opportunities:
        d = dict(r)
        d["suggested_price"] = round(float(d["unit_cost"] or 0) * 1.10, 2)
        try:
            oldest = date.fromisoformat(d["oldest_date"])
            d["days_in_stock"] = (today - oldest).days
        except Exception:
            d["days_in_stock"] = 0
        opp.append(d)

    return render_template(
        "dashboard.html",
        sales_today=sales_today,
        sales_today_count=sales_today_count,
        sales_month=sales_month,
        payments_month=payments_month,
        stock_available=stock_available,
        top_products=top_products,
        opportunities=opp,
        chassis_alerts=chassis_alerts,
    )


@app.route("/products", methods=["GET", "POST"])
@login_required
@roles_required("admin", "finance", "stock", "support")
def products():
    if request.method == "POST":
        name = request.form["name"].strip()
        sku = request.form.get("sku", "").strip() or None
        category = request.form.get("category", "").strip()
        fob_price_usd = float(request.form.get("fob_price_usd") or 0)
        aliquota_rate = float(request.form.get("aliquota_rate") or 0)
        unit_cost = float(request.form.get("unit_cost") or 0)

        if fob_price_usd > 0 and aliquota_rate > 0:
            unit_cost = round(fob_price_usd * aliquota_rate, 2)
        elif unit_cost > 0 and fob_price_usd > 0 and aliquota_rate == 0:
            aliquota_rate = round(unit_cost / fob_price_usd, 4)
        elif unit_cost > 0 and aliquota_rate > 0 and fob_price_usd == 0:
            fob_price_usd = round(unit_cost / aliquota_rate, 2)

        retail_price = float(request.form.get("retail_price") or 0)
        wholesale_price = float(request.form.get("wholesale_price") or 0)
        installment_12x = float(request.form.get("installment_12x") or 0)
        installment_18x = float(request.form.get("installment_18x") or 0)
        if retail_price > 0:
            if installment_12x == 0:
                installment_12x = round((retail_price * 1.1013216) / 12, 2)
            if installment_18x == 0:
                installment_18x = round((retail_price * 1.1437722) / 18, 2)
        promo_eligible = 1 if request.form.get("promo_eligible") else 0
        with db() as conn:
            conn.execute(
                """INSERT INTO products(name,sku,category,fob_price_usd,aliquota_rate,unit_cost,retail_price,wholesale_price,installment_12x,installment_18x,promo_eligible)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (name, sku, category, fob_price_usd, aliquota_rate, unit_cost, retail_price, wholesale_price, installment_12x, installment_18x, promo_eligible),
            )
            conn.commit()
        audit("product.created", name)
        flash("Produto cadastrado.", "success")
        return redirect(url_for("products"))
    with db() as conn:
        rows = conn.execute(
            """
            SELECT p.*,
                   COALESCE(p.bling_stock, 0) as bling_stock,
                   p.bling_updated_at,
                   SUM(CASE WHEN st.status='available' THEN 1 ELSE 0 END) local_available,
                   SUM(CASE WHEN st.status='sold' THEN 1 ELSE 0 END) sold
            FROM products p LEFT JOIN stock_units st ON st.product_id=p.id
            GROUP BY p.id ORDER BY p.name
            """
        ).fetchall()
        
        # Determine last sync timestamp
        last_sync = None
        for r in rows:
            if r["bling_updated_at"]:
                last_sync = r["bling_updated_at"]
                break

    return render_template("products.html", products=rows, last_sync=last_sync)


@app.route("/products/<int:pid>/edit", methods=["POST"])
@login_required
@roles_required("admin", "finance", "support")
def edit_product(pid):
    retail_price = float(request.form.get("retail_price") or 0)
    installment_12x = float(request.form.get("installment_12x") or 0)
    installment_18x = float(request.form.get("installment_18x") or 0)
    if retail_price > 0:
        if installment_12x == 0:
            installment_12x = round((retail_price * 1.1013216) / 12, 2)
        if installment_18x == 0:
            installment_18x = round((retail_price * 1.1437722) / 18, 2)

    fob_price_usd = float(request.form.get("fob_price_usd") or 0)
    aliquota_rate = float(request.form.get("aliquota_rate") or 0)
    unit_cost = float(request.form.get("unit_cost") or 0)

    if fob_price_usd > 0 and aliquota_rate > 0:
        unit_cost = round(fob_price_usd * aliquota_rate, 2)
    elif unit_cost > 0 and fob_price_usd > 0 and aliquota_rate == 0:
        aliquota_rate = round(unit_cost / fob_price_usd, 4)
    elif unit_cost > 0 and aliquota_rate > 0 and fob_price_usd == 0:
        fob_price_usd = round(unit_cost / aliquota_rate, 2)

    fields = (
        request.form.get("name", "").strip(),
        request.form.get("sku", "").strip() or None,
        request.form.get("category", "").strip(),
        fob_price_usd,
        aliquota_rate,
        unit_cost,
        retail_price,
        float(request.form.get("wholesale_price") or 0),
        installment_12x,
        installment_18x,
        1 if request.form.get("promo_eligible") else 0,
        pid,
    )
    with db() as conn:
        conn.execute(
            """UPDATE products SET name=?,sku=?,category=?,fob_price_usd=?,aliquota_rate=?,unit_cost=?,retail_price=?,wholesale_price=?,installment_12x=?,installment_18x=?,promo_eligible=?
               WHERE id=?""",
            fields,
        )
        conn.commit()
    audit("product.updated", f"product_id={pid}")
    flash("Produto atualizado.", "success")
    return redirect(url_for("products"))


def clean_product_name(raw_name: str) -> str:
    return (raw_name or "").strip()


def parse_products_rows(data_bytes=None, text_content=None, filename="sheet.csv"):
    rows = []
    all_rows = []
    if text_content:
        text = text_content
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"
        reader = csv.reader(io.StringIO(text), dialect)
        all_rows = list(reader)
    elif data_bytes:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "csv"
        if ext == "xlsx":
            if load_workbook is None:
                raise ValueError("Suporte a XLSX indisponível. Instale openpyxl.")
            wb = load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
            ws = wb.active
            all_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        else:
            text = data_bytes.decode("utf-8-sig", errors="replace")
            sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except Exception:
                dialect = csv.excel
                dialect.delimiter = ";"
            reader = csv.reader(io.StringIO(text), dialect)
            all_rows = list(reader)

    if not all_rows:
        return []

    h_idx, headers, data_rows = find_header_and_data_rows(all_rows)
    def idx(candidates):
        for c in candidates:
            if c in headers:
                return headers.index(c)
        return None

    i_name = idx(["produto", "product", "modelo", "model", "nome", "name", "descricao", "description"])
    i_sku = idx(["sku", "codigo", "code"])
    i_category = idx(["categoria", "category", "tipo", "type"])
    i_fob = idx(["fob", "fob (usd)", "fob usd", "preco fob", "usd fob", "valor fob"])
    i_aliquota = idx(["aliquota", "aliquota real", "aliquota (r$/usd)", "taxa aliquota", "aliquota r$"])
    i_cost = idx(["custo", "unit_cost", "cost", "custo unitario", "valor custo", "custo galpao"])
    i_wholesale = idx(["atacado", "wholesale", "wholesale_price", "preco atacado", "valor atacado"])
    i_retail = idx(["varejo", "retail", "retail_price", "preco varejo", "valor varejo", "preco", "price"])
    i_inst_12x = idx(["12x varejo", "12x", "12x (parcela)", "parcela 12x", "12x varejo (parcela)"])
    i_inst_18x = idx(["18x varejo", "18x", "18x (parcela)", "parcela 18x", "18x varejo (parcela)"])
    i_promo = idx(["promo", "promo_eligible", "elegivel", "promocional"])

    if i_name is None and i_sku is None:
        raise ValueError("A planilha precisa conter ao menos a coluna 'PRODUTO', 'MODELO' ou 'SKU'.")

    def parse_float(val):
        if val is None:
            return 0.0
        s = str(val).strip().replace("R$", "").replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            try:
                return float(str(val).strip().replace(",", "."))
            except Exception:
                return 0.0

    for raw in data_rows:
        name = str(raw[i_name] or "").strip() if i_name is not None and i_name < len(raw) else ""
        sku = str(raw[i_sku] or "").strip() if i_sku is not None and i_sku < len(raw) else ""
        category = str(raw[i_category] or "").strip() if i_category is not None and i_category < len(raw) else ""
        fob_price_usd = parse_float(raw[i_fob]) if i_fob is not None and i_fob < len(raw) else 0.0
        aliquota_rate = parse_float(raw[i_aliquota]) if i_aliquota is not None and i_aliquota < len(raw) else 0.0
        unit_cost = parse_float(raw[i_cost]) if i_cost is not None and i_cost < len(raw) else 0.0

        if fob_price_usd > 0 and aliquota_rate > 0:
            unit_cost = round(fob_price_usd * aliquota_rate, 2)
        elif unit_cost > 0 and fob_price_usd > 0 and aliquota_rate == 0:
            aliquota_rate = round(unit_cost / fob_price_usd, 4)
        elif unit_cost > 0 and aliquota_rate > 0 and fob_price_usd == 0:
            fob_price_usd = round(unit_cost / aliquota_rate, 2)

        wholesale_price = parse_float(raw[i_wholesale]) if i_wholesale is not None and i_wholesale < len(raw) else 0.0
        retail_price = parse_float(raw[i_retail]) if i_retail is not None and i_retail < len(raw) else 0.0
        installment_12x = parse_float(raw[i_inst_12x]) if i_inst_12x is not None and i_inst_12x < len(raw) else 0.0
        installment_18x = parse_float(raw[i_inst_18x]) if i_inst_18x is not None and i_inst_18x < len(raw) else 0.0
        if retail_price > 0:
            if installment_12x == 0:
                installment_12x = round((retail_price * 1.1013216) / 12, 2)
            if installment_18x == 0:
                installment_18x = round((retail_price * 1.1437722) / 18, 2)
        
        promo_val = str(raw[i_promo] or "").strip().lower() if i_promo is not None and i_promo < len(raw) else "1"
        promo_eligible = True if promo_val in ["1", "true", "sim", "s", "elegivel", "yes"] else False

        if name or sku:
            rows.append({
                "name": clean_product_name(name) if name else sku,
                "sku": sku or None,
                "category": category or None,
                "fob_price_usd": fob_price_usd,
                "aliquota_rate": aliquota_rate,
                "unit_cost": unit_cost,
                "wholesale_price": wholesale_price,
                "retail_price": retail_price,
                "installment_12x": installment_12x,
                "installment_18x": installment_18x,
                "promo_eligible": promo_eligible
            })
    return rows


@app.route("/products/import", methods=["POST"])
@login_required
@roles_required("admin", "finance", "stock")
def import_products():
    import urllib.request
    sheets_url = request.form.get("sheets_url", "").strip()
    file_storage = request.files.get("products_file")

    parsed_rows = []

    if sheets_url:
        export_url = sheets_url
        if "docs.google.com/spreadsheets" in export_url and "/export" not in export_url:
            export_url = export_url.split("/edit")[0].rstrip("/") + "/export?format=csv"
            gid = None
            if "#gid=" in sheets_url:
                gid = sheets_url.split("#gid=")[1].split("&")[0]
            elif "gid=" in sheets_url:
                gid = sheets_url.split("gid=")[1].split("&")[0]
            if gid:
                export_url += f"&gid={gid}"

        try:
            req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                csv_text = resp.read().decode("utf-8-sig", errors="replace")
                parsed_rows = parse_products_rows(text_content=csv_text)
        except Exception as e:
            flash(f"Falha ao carregar a planilha do Google Sheets pelo link. Verifique se a permissão do link está como 'Qualquer pessoa com o link pode ver'. Erro: {str(e)}", "danger")
            return redirect(url_for("products"))

    elif file_storage and file_storage.filename:
        data = file_storage.read()
        try:
            parsed_rows = parse_products_rows(data_bytes=data, filename=file_storage.filename)
        except Exception as e:
            flash(f"Erro ao processar arquivo: {str(e)}", "danger")
            return redirect(url_for("products"))
    else:
        flash("Selecione um arquivo CSV/XLSX ou informe a URL do Google Sheets.", "danger")
        return redirect(url_for("products"))

    if not parsed_rows:
        flash("Nenhum produto válido encontrado na planilha.", "warning")
        return redirect(url_for("products"))

    created_count = 0
    updated_count = 0

    with db() as conn:
        for r in parsed_rows:
            existing = None
            if r["sku"]:
                existing = conn.execute("SELECT id FROM products WHERE lower(sku)=lower(?)", (r["sku"],)).fetchone()
            if not existing and r["name"]:
                existing = conn.execute("SELECT id FROM products WHERE lower(name)=lower(?)", (r["name"],)).fetchone()

            if existing:
                conn.execute(
                    """UPDATE products SET name=?, sku=COALESCE(?, sku), category=COALESCE(?, category),
                       fob_price_usd=?, aliquota_rate=?, unit_cost=?, wholesale_price=?, retail_price=?, installment_12x=?, installment_18x=?, promo_eligible=? WHERE id=?""",
                    (r["name"], r["sku"], r["category"], r.get("fob_price_usd", 0), r.get("aliquota_rate", 0), r["unit_cost"], r["wholesale_price"], r["retail_price"], r.get("installment_12x", 0), r.get("installment_18x", 0), r["promo_eligible"], existing["id"])
                )
                updated_count += 1
            else:
                conn.execute(
                    """INSERT INTO products (name, sku, category, fob_price_usd, aliquota_rate, unit_cost, wholesale_price, retail_price, installment_12x, installment_18x, promo_eligible)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (r["name"], r["sku"], r["category"], r.get("fob_price_usd", 0), r.get("aliquota_rate", 0), r["unit_cost"], r["wholesale_price"], r["retail_price"], r.get("installment_12x", 0), r.get("installment_18x", 0), r["promo_eligible"])
                )
                created_count += 1
        conn.commit()
        conn.commit()

    audit("products.imported", f"created={created_count}; updated={updated_count}")
    flash(f"Planilha de produtos processada! {created_count} novos produtos cadastrados, {updated_count} atualizados.", "success")
    return redirect(url_for("products"))


@app.route("/api/sync-prices", methods=["POST"])
@app.route("/sync-prices", methods=["POST"])
def api_sync_prices():

    token = request.args.get("token") or request.headers.get("X-Sync-Token")
    expected_token = os.environ.get("SYNC_TOKEN") or "maj-m-one-sync-secret-2026"
    if not token or token != expected_token:
        return jsonify({"ok": False, "error": "Token de autenticação inválido"}), 401

    payload = request.get_json(silent=True)
    all_rows = []

    if isinstance(payload, list):
        all_rows = payload
    elif isinstance(payload, dict):
        all_rows = payload.get("data") or payload.get("rows") or payload.get("products") or []

    if not all_rows and request.data:
        text = request.data.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"
        reader = csv.reader(io.StringIO(text), dialect)
        all_rows = list(reader)

    if not all_rows:
        return jsonify({"ok": False, "error": "Nenhum dado enviado"}), 400

    parsed_rows = []
    if isinstance(all_rows[0], list):
        h_idx, headers, data_rows = find_header_and_data_rows(all_rows)
        def idx(candidates):
            for c in candidates:
                if c in headers:
                    return headers.index(c)
            return None

        i_name = idx(["produto", "product", "modelo", "model", "nome", "name", "descricao", "description"])
        i_sku = idx(["sku", "codigo", "code"])
        i_category = idx(["categoria", "category", "tipo", "type"])
        i_cost = idx(["custo", "unit_cost", "cost", "custo unitario", "valor custo"])
        i_wholesale = idx(["atacado", "wholesale", "wholesale_price", "preco atacado", "valor atacado"])
        i_retail = idx(["varejo", "retail", "retail_price", "preco varejo", "valor varejo", "preco", "price"])
        i_inst_12x = idx(["12x varejo", "12x", "12x (parcela)", "parcela 12x", "12x varejo (parcela)"])
        i_inst_18x = idx(["18x varejo", "18x", "18x (parcela)", "parcela 18x", "18x varejo (parcela)"])
        i_promo = idx(["promo", "promo_eligible", "elegivel", "promocional"])

        def parse_float(val):
            if val is None:
                return 0.0
            s = str(val).strip().replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
            try:
                return float(s)
            except Exception:
                try:
                    return float(str(val).strip().replace(",", "."))
                except Exception:
                    return 0.0

        for raw in data_rows:

            name = str(raw[i_name] or "").strip() if i_name is not None and i_name < len(raw) else ""
            sku = str(raw[i_sku] or "").strip() if i_sku is not None and i_sku < len(raw) else ""
            category = str(raw[i_category] or "").strip() if i_category is not None and i_category < len(raw) else ""
            unit_cost = parse_float(raw[i_cost]) if i_cost is not None and i_cost < len(raw) else 0.0
            wholesale_price = parse_float(raw[i_wholesale]) if i_wholesale is not None and i_wholesale < len(raw) else 0.0
            retail_price = parse_float(raw[i_retail]) if i_retail is not None and i_retail < len(raw) else 0.0
            installment_12x = parse_float(raw[i_inst_12x]) if i_inst_12x is not None and i_inst_12x < len(raw) else 0.0
            installment_18x = parse_float(raw[i_inst_18x]) if i_inst_18x is not None and i_inst_18x < len(raw) else 0.0
            
            promo_val = str(raw[i_promo] or "").strip().lower() if i_promo is not None and i_promo < len(raw) else "1"
            promo_eligible = True if promo_val in ["1", "true", "sim", "s", "elegivel", "yes"] else False

            if name or sku:
                parsed_rows.append({
                    "name": clean_product_name(name) if name else sku,
                    "sku": sku or None,
                    "category": category or None,
                    "unit_cost": unit_cost,
                    "wholesale_price": wholesale_price,
                    "retail_price": retail_price,
                    "installment_12x": installment_12x,
                    "installment_18x": installment_18x,
                    "promo_eligible": promo_eligible
                })
    elif isinstance(all_rows[0], dict):
        for item in all_rows:
            name = str(item.get("name") or item.get("produto") or item.get("modelo") or "").strip()
            sku = str(item.get("sku") or item.get("codigo") or "").strip() or None
            category = str(item.get("category") or item.get("categoria") or "").strip() or None
            unit_cost = float(item.get("unit_cost") or item.get("custo") or 0)
            wholesale_price = float(item.get("wholesale_price") or item.get("atacado") or 0)
            retail_price = float(item.get("retail_price") or item.get("varejo") or 0)
            installment_12x = float(item.get("installment_12x") or item.get("12x") or 0)
            installment_18x = float(item.get("installment_18x") or item.get("18x") or 0)
            promo_eligible = bool(item.get("promo_eligible", True))
            if name or sku:
                parsed_rows.append({
                    "name": clean_product_name(name) if name else sku,
                    "sku": sku,
                    "category": category,
                    "unit_cost": unit_cost,
                    "wholesale_price": wholesale_price,
                    "retail_price": retail_price,
                    "installment_12x": installment_12x,
                    "installment_18x": installment_18x,
                    "promo_eligible": promo_eligible
                })


    created_count = 0
    updated_count = 0

    with db() as conn:
        for r in parsed_rows:
            existing = None
            if r["sku"]:
                existing = conn.execute("SELECT id FROM products WHERE lower(sku)=lower(?)", (r["sku"],)).fetchone()
            if not existing and r["name"]:
                existing = conn.execute("SELECT id FROM products WHERE lower(name)=lower(?)", (r["name"],)).fetchone()

            if existing:
                conn.execute(
                    """UPDATE products SET name=?, sku=COALESCE(?, sku), category=COALESCE(?, category),
                       unit_cost=?, wholesale_price=?, retail_price=?, installment_12x=?, installment_18x=?, promo_eligible=? WHERE id=?""",
                    (r["name"], r["sku"], r["category"], r["unit_cost"], r["wholesale_price"], r["retail_price"], r.get("installment_12x", 0), r.get("installment_18x", 0), r["promo_eligible"], existing["id"])
                )
                updated_count += 1
            else:
                conn.execute(
                    """INSERT INTO products (name, sku, category, unit_cost, wholesale_price, retail_price, installment_12x, installment_18x, promo_eligible)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (r["name"], r["sku"], r["category"], r["unit_cost"], r["wholesale_price"], r["retail_price"], r.get("installment_12x", 0), r.get("installment_18x", 0), r["promo_eligible"])
                )
                created_count += 1
        conn.commit()

    audit("products.sync_webhook", f"created={created_count}; updated={updated_count}")
    return jsonify({
        "ok": True,
        "message": f"Sincronização concluída com sucesso. {created_count} criados, {updated_count} atualizados.",
        "created": created_count,
        "updated": updated_count
    })




@app.route("/imports", methods=["GET", "POST"])
@login_required
@roles_required("admin", "support")
def imports():
    if request.method == "POST":
        reference = request.form["reference"].strip()
        invoice_no = request.form.get("invoice_no", "").strip()
        bl_no = request.form.get("bl_no", "").strip()
        supplier_name = request.form.get("supplier_name", "").strip()
        seller_name = request.form.get("seller_name", "").strip()
        arrival_date = request.form.get("arrival_date") or None
        usd_rate = float(request.form.get("usd_rate") or 0)
        invoice_amount_usd = float(request.form.get("invoice_amount_usd") or 0)
        nf_entry = request.form.get("nf_entry", "").strip()
        notes = request.form.get("notes", "").strip()
        try:
            invoice_file = save_upload(request.files.get("invoice_file"), "invoice")
            bl_file = save_upload(request.files.get("bl_file"), "bl")
            nf_entry_file = save_upload(request.files.get("nf_entry_file"), "nfentrada")
            chassis_file_obj = request.files.get("chassis_file")
            chassis_file_name = save_upload(chassis_file_obj, "chassis") if chassis_file_obj and chassis_file_obj.filename else None
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("imports"))

        with db() as conn:
            cur = conn.execute(
                """INSERT INTO imports(reference,invoice_no,bl_no,supplier_name,seller_name,arrival_date,usd_rate,invoice_amount_usd,nf_entry,invoice_file,bl_file,nf_entry_file,chassis_file,notes,created_by)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (reference, invoice_no, bl_no, supplier_name, seller_name, arrival_date, usd_rate, invoice_amount_usd, nf_entry, invoice_file, bl_file, nf_entry_file, chassis_file_name, notes, session["user_id"]),
            )
            iid = cur.lastrowid

            if chassis_file_name:
                try:
                    class FS:
                        pass
                    obj = FS()
                    obj.filename = chassis_file_name
                    obj.read = lambda: (UPLOAD_DIR / chassis_file_name).read_bytes()
                    rows = parse_chassis_file(obj)
                    inserted = 0
                    for row in rows:
                        existing = conn.execute("SELECT id FROM stock_units WHERE chassis=?", (row["chassis"],)).fetchone()
                        if not existing:
                            prod = conn.execute("SELECT id FROM products WHERE lower(name)=lower(?)", (row["model"],)).fetchone()
                            if not prod:
                                sku_base = "".join(ch for ch in row["model"].upper() if ch.isalnum())[:18] or "PROD"
                                sku = sku_base
                                n = 1
                                while conn.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone():
                                    n += 1
                                    sku = f"{sku_base}-{n}"
                                cur_p = conn.execute("INSERT INTO products(name,sku,category) VALUES(?,?,?)", (row["model"], sku, "Importado"))
                                product_id = cur_p.lastrowid
                            else:
                                product_id = prod["id"]
                            conn.execute(
                                "INSERT INTO stock_units(chassis,motor_no,product_id,color,import_id,status,received_at) VALUES(?,?,?,?,?,?,?)",
                                (row["chassis"], row["motor"], product_id, row["color"], iid, "unreleased", arrival_date),
                            )
                            inserted += 1
                    flash(f"Importação criada com {inserted} chassis cadastrados.", "success")
                except Exception as ex:
                    flash(f"Importação criada, porém erro ao processar planilha de chassis: {ex}", "warning")
            else:
                flash("Importação criada com sucesso. Carregue a planilha de chassis para liberar o estoque.", "success")
            conn.commit()

        audit("import.created", reference)
        return redirect(url_for("imports"))

    with db() as conn:
        import_rows = conn.execute(
            """
            SELECT i.*, COUNT(DISTINCT st.id) AS chassis_count
            FROM imports i
            LEFT JOIN stock_units st ON st.import_id=i.id
            GROUP BY i.id ORDER BY i.created_at DESC
            """
        ).fetchall()

        result_imports = []
        for row in import_rows:
            imp_dict = dict(row)
            costs = conn.execute(
                "SELECT * FROM import_costs WHERE import_id=? ORDER BY paid_at ASC, id ASC",
                (imp_dict["id"],)
            ).fetchall()

            costs_list = [dict(c) for c in costs]
            supplier_brl = 0.0
            supplier_usd = 0.0
            total_costs_brl = 0.0

            for c in costs_list:
                c_rate = float(c.get("usd_rate") or imp_dict.get("usd_rate") or 1.0)
                amount = float(c.get("amount") or 0.0)
                c_currency = (c.get("currency") or "BRL").upper()

                if c_currency == "USD":
                    amount_usd = amount
                    amount_brl = amount * (c_rate if c_rate > 0 else 1.0)
                else:
                    amount_brl = amount
                    amount_usd = amount / c_rate if c_rate > 0 else 0.0

                c["amount_brl"] = amount_brl
                c["amount_usd"] = amount_usd
                total_costs_brl += amount_brl

                c_type = (c.get("cost_type") or "").strip().lower()
                c_desc = (c.get("description") or "").strip().lower()
                is_supplier_payment = any(kw in c_type or kw in c_desc for kw in ["fornecedor", "sinal", "inicial", "intermedi", "final", "invoice", "china", "chines", "fabricante"])

                if is_supplier_payment or c_currency == "USD":
                    supplier_brl += amount_brl
                    if c_currency == "USD":
                        supplier_usd += amount
                    elif c_rate > 0:
                        supplier_usd += (amount / c_rate)

            imp_dict["costs_list"] = costs_list
            imp_dict["costs_brl"] = total_costs_brl
            imp_dict["supplier_brl"] = supplier_brl
            imp_dict["supplier_usd"] = supplier_usd

            # Cálculo do Câmbio Médio (Dólar Médio dos pagamentos chineses)
            if supplier_usd > 0 and supplier_brl > 0:
                imp_dict["avg_usd_rate"] = round(supplier_brl / supplier_usd, 4)
            elif float(imp_dict.get("invoice_amount_usd") or 0) > 0 and supplier_brl > 0:
                imp_dict["avg_usd_rate"] = round(supplier_brl / float(imp_dict.get("invoice_amount_usd")), 4)
            else:
                imp_dict["avg_usd_rate"] = float(imp_dict.get("usd_rate") or 0.0)

            # Cálculo da Alíquota Real do Dólar da Importação (R$ Total de Despesas / US$ Invoice)
            inv_usd = float(imp_dict.get("invoice_amount_usd") or 0.0)
            if inv_usd > 0 and total_costs_brl > 0:
                imp_dict["landed_aliquota"] = round(total_costs_brl / inv_usd, 4)
            elif supplier_usd > 0 and total_costs_brl > 0:
                imp_dict["landed_aliquota"] = round(total_costs_brl / supplier_usd, 4)
            else:
                imp_dict["landed_aliquota"] = 0.0

            result_imports.append(imp_dict)

    return render_template("imports.html", imports=result_imports)


@app.route("/imports/<int:iid>/edit", methods=["POST"])
@login_required
@roles_required("admin")
def edit_import(iid):
    reference = request.form.get("reference", "").strip()
    invoice_no = request.form.get("invoice_no", "").strip()
    bl_no = request.form.get("bl_no", "").strip()
    supplier_name = request.form.get("supplier_name", "").strip()
    seller_name = request.form.get("seller_name", "").strip()
    nf_entry = request.form.get("nf_entry", "").strip()
    arrival_date = request.form.get("arrival_date") or None
    usd_rate = float(request.form.get("usd_rate") or 0)
    invoice_amount_usd = float(request.form.get("invoice_amount_usd") or 0)
    notes = request.form.get("notes", "").strip()

    with db() as conn:
        imp = conn.execute("SELECT * FROM imports WHERE id=?", (iid,)).fetchone()
        if not imp:
            flash("Importação não encontrada.", "danger")
            return redirect(url_for("imports"))

        try:
            invoice_file = save_upload(request.files.get("invoice_file"), "invoice") or imp["invoice_file"]
            bl_file = save_upload(request.files.get("bl_file"), "bl") or imp["bl_file"]
            nf_entry_file = save_upload(request.files.get("nf_entry_file"), "nfentrada") or imp["nf_entry_file"]
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("imports"))

        chassis_file = imp["chassis_file"]
        chassis_file_obj = request.files.get("chassis_file")
        if chassis_file_obj and chassis_file_obj.filename:
            try:
                chassis_file = save_upload(chassis_file_obj, "chassis")
                class FS:
                    pass
                obj = FS()
                obj.filename = chassis_file
                obj.read = lambda: (UPLOAD_DIR / chassis_file).read_bytes()
                rows = parse_chassis_file(obj)
                inserted = 0
                for row in rows:
                    existing = conn.execute("SELECT id FROM stock_units WHERE chassis=?", (row["chassis"],)).fetchone()
                    if not existing:
                        prod = conn.execute("SELECT id FROM products WHERE lower(name)=lower(?)", (row["model"],)).fetchone()
                        if not prod:
                            sku_base = "".join(ch for ch in row["model"].upper() if ch.isalnum())[:18] or "PROD"
                            sku = sku_base
                            n = 1
                            while conn.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone():
                                n += 1
                                sku = f"{sku_base}-{n}"
                            cur_p = conn.execute("INSERT INTO products(name,sku,category) VALUES(?,?,?)", (row["model"], sku, "Importado"))
                            product_id = cur_p.lastrowid
                        else:
                            product_id = prod["id"]
                        conn.execute(
                            "INSERT INTO stock_units(chassis,motor_no,product_id,color,import_id,status,received_at) VALUES(?,?,?,?,?,?,?)",
                            (row["chassis"], row["motor"], product_id, row["color"], iid, "available" if imp["status"] == "released" else "unreleased", arrival_date),
                        )
                        inserted += 1
                flash(f"Planilha de chassis atualizada ({inserted} novos chassis).", "info")
            except Exception as ex:
                flash(f"Erro ao ler planilha de chassis: {ex}", "warning")

        conn.execute(
            """UPDATE imports SET reference=?, invoice_no=?, bl_no=?, supplier_name=?, seller_name=?, nf_entry=?, arrival_date=?, usd_rate=?, invoice_amount_usd=?, invoice_file=?, bl_file=?, nf_entry_file=?, chassis_file=?, notes=?
               WHERE id=?""",
            (reference or imp["reference"], invoice_no, bl_no, supplier_name, seller_name, nf_entry, arrival_date, usd_rate, invoice_amount_usd, invoice_file, bl_file, nf_entry_file, chassis_file, notes, iid)
        )
        conn.commit()
    audit("import.updated", f"import_id={iid}")
    flash("Importação atualizada com sucesso.", "success")
    return redirect(url_for("imports"))


def normalize_headers(headers):
    out = []
    for h in headers:
        s = str(h or "").strip().lower()
        s = s.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
        out.append(s)
    return out


def parse_chassis_file(file_storage):
    ext = file_storage.filename.rsplit(".", 1)[1].lower()
    data = file_storage.read()
    rows = []
    if ext == "csv":
        text = data.decode("utf-8-sig", errors="replace")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"
        reader = csv.reader(io.StringIO(text), dialect)
        all_rows = list(reader)
    elif ext == "xlsx":
        if load_workbook is None:
            raise ValueError("Suporte a XLSX indisponível. Instale openpyxl.")
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        all_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    else:
        raise ValueError("Use CSV ou XLSX para a planilha de chassis.")
    if not all_rows:
        return []
    headers = normalize_headers(all_rows[0])
    def idx(candidates):
        for c in candidates:
            if c in headers:
                return headers.index(c)
        return None
    i_model = idx(["modelo", "model", "produto", "product"])
    i_chassis = idx(["chassi", "chassis", "quadro", "frame", "frame no", "frame number", "vin"])
    i_motor = idx(["motor", "motor no", "motor number", "numero do motor", "n motor"])
    i_color = idx(["cor", "color", "colour"])
    if i_model is None or i_chassis is None:
        raise ValueError("A planilha precisa ter pelo menos as colunas MODELO e CHASSI.")
    for raw in all_rows[1:]:
        model = str(raw[i_model] or "").strip() if i_model < len(raw) else ""
        chassis = str(raw[i_chassis] or "").strip() if i_chassis < len(raw) else ""
        motor = str(raw[i_motor] or "").strip() if i_motor is not None and i_motor < len(raw) else ""
        color = str(raw[i_color] or "").strip() if i_color is not None and i_color < len(raw) else ""
        if model and chassis:
            rows.append({"model": model, "chassis": chassis, "motor": motor, "color": color})
    return rows


@app.route("/imports/<int:iid>/chassis", methods=["POST"])
@login_required
@roles_required("admin", "stock", "support")
def import_chassis(iid):
    f = request.files.get("chassis_file")
    if not f or not f.filename:
        flash("Selecione uma planilha CSV ou XLSX.", "danger")
        return redirect(url_for("imports"))
    try:
        filename = save_upload(f, "chassis")
        # Re-open saved file for parsing.
        class FS:
            pass
        obj = FS()
        obj.filename = filename
        obj.read = lambda: (UPLOAD_DIR / filename).read_bytes()
        rows = parse_chassis_file(obj)
    except Exception as e:
        flash(f"Não foi possível importar a planilha: {e}", "danger")
        return redirect(url_for("imports"))
    inserted = 0
    duplicates = []
    with db() as conn:
        imp = conn.execute("SELECT * FROM imports WHERE id=?", (iid,)).fetchone()
        if not imp:
            flash("Importação não encontrada.", "danger")
            return redirect(url_for("imports"))
        for row in rows:
            existing = conn.execute("SELECT id FROM stock_units WHERE chassis=?", (row["chassis"],)).fetchone()
            if existing:
                duplicates.append(row["chassis"])
                continue
            prod = conn.execute("SELECT id FROM products WHERE lower(name)=lower(?)", (row["model"],)).fetchone()
            if not prod:
                sku_base = "".join(ch for ch in row["model"].upper() if ch.isalnum())[:18] or "PROD"
                sku = sku_base
                n = 1
                while conn.execute("SELECT 1 FROM products WHERE sku=?", (sku,)).fetchone():
                    n += 1
                    sku = f"{sku_base}-{n}"
                cur = conn.execute("INSERT INTO products(name,sku,category) VALUES(?,?,?)", (row["model"], sku, "Importado"))
                product_id = cur.lastrowid
            else:
                product_id = prod["id"]
            conn.execute(
                "INSERT INTO stock_units(chassis,motor_no,product_id,color,import_id,status,received_at) VALUES(?,?,?,?,?,?,?)",
                (row["chassis"], row["motor"], product_id, row["color"], iid, "available" if imp["status"] == "released" else "unreleased", imp["arrival_date"]),
            )
            inserted += 1
        conn.execute("UPDATE imports SET chassis_file=? WHERE id=?", (filename, iid))
        conn.commit()
    audit("import.chassis", f"import_id={iid}; inserted={inserted}; duplicates={len(duplicates)}")
    msg = f"{inserted} chassis importados."
    if duplicates:
        msg += f" {len(duplicates)} duplicados foram bloqueados."
    flash(msg, "success" if inserted else "warning")
    return redirect(url_for("imports"))


@app.route("/imports/<int:iid>/cost", methods=["POST"])
@login_required
@roles_required("admin", "support")
def add_import_cost(iid):
    try:
        receipt = save_upload(request.files.get("receipt_file"), "importcost")
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("imports"))
    with db() as conn:
        conn.execute(
            "INSERT INTO import_costs(import_id,cost_type,description,amount,currency,usd_rate,paid_at,receipt_file) VALUES(?,?,?,?,?,?,?,?)",
            (
                iid,
                request.form.get("cost_type", "Pagamento Extra / Outros"),
                request.form.get("description", "").strip(),
                float(request.form.get("amount") or 0),
                request.form.get("currency", "BRL"),
                float(request.form.get("usd_rate") or 0),
                request.form.get("paid_at") or date.today().isoformat(),
                receipt,
            ),
        )
        conn.commit()
    audit("import.cost", f"import_id={iid}")
    flash("Comprovante/Custo adicionado à importação.", "success")
    return redirect(url_for("imports"))


@app.route("/imports/cost/<int:cid>/delete", methods=["POST"])
@login_required
@roles_required("admin")
def delete_import_cost(cid):
    with db() as conn:
        cost = conn.execute("SELECT import_id FROM import_costs WHERE id=?", (cid,)).fetchone()
        if cost:
            conn.execute("DELETE FROM import_costs WHERE id=?", (cid,))
            conn.commit()
            audit("import.cost_deleted", f"cost_id={cid}")
            flash("Comprovante/Custo removido com sucesso.", "success")
    return redirect(url_for("imports"))


@app.route("/imports/<int:iid>/release", methods=["POST"])
@login_required
@roles_required("admin", "support")
def release_import(iid):
    with db() as conn:
        imp = conn.execute("SELECT * FROM imports WHERE id=?", (iid,)).fetchone()
        count = conn.execute("SELECT COUNT(*) c FROM stock_units WHERE import_id=?", (iid,)).fetchone()["c"]
        if not imp:
            flash("Importação não encontrada.", "danger")
        elif count == 0:
            flash("BLOQUEIO DE SEGURANÇA: Não é possível liberar a importação para venda sem cadastrar os chassis da remessa. Carregue a planilha de chassis primeiro.", "danger")
        elif not imp["invoice_no"] or not imp["bl_no"]:
            flash("Para liberar a importação, é necessário informar Invoice e BL.", "danger")
        else:
            conn.execute("UPDATE imports SET status='released' WHERE id=?", (iid,))
            conn.execute("UPDATE stock_units SET status='available' WHERE import_id=? AND status='unreleased'", (iid,))
            conn.commit()
            audit("import.released", f"import_id={iid}")
            flash("Estoque desta importação foi liberado com sucesso para venda.", "success")
    return redirect(url_for("imports"))


@app.route("/api/imports/analyze-docs", methods=["POST"])
@login_required
@roles_required("admin", "support")
def api_analyze_import_docs():
    file_objs = []
    for key in ["invoice_file", "bl_file", "nf_entry_file", "chassis_file"]:
        f = request.files.get(key)
        if f and f.filename:
            content = f.read()
            ext = f.filename.rsplit(".", 1)[-1].lower()
            mime = "application/pdf" if ext == "pdf" else ("text/csv" if ext == "csv" else f"image/{ext if ext != 'jpg' else 'jpeg'}")
            file_objs.append({
                "bytes": content,
                "mime_type": mime,
                "filename": f.filename
            })

    if not file_objs:
        return jsonify({"success": False, "message": "Nenhum arquivo enviado para análise da IA."})

    res = analyze_import_documents(file_objs)
    return jsonify(res)


@app.route("/stock")
@login_required
def stock():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "available")
    params = []
    where = ["1=1"]
    if q:
        where.append("(st.chassis LIKE ? OR p.name LIKE ? OR st.color LIKE ? OR st.motor_no LIKE ?)")
        term = f"%{q}%"
        params += [term, term, term, term]
    if status and status != "all":
        where.append("st.status=?")
        params.append(status)
    with db() as conn:
        rows = conn.execute(
            f"""
            SELECT st.*,p.name product_name,i.reference import_ref,i.status import_status,s.invoice_number
            FROM stock_units st JOIN products p ON p.id=st.product_id
            LEFT JOIN imports i ON i.id=st.import_id
            LEFT JOIN sales s ON s.id=st.sale_id
            WHERE {' AND '.join(where)} ORDER BY st.created_at DESC LIMIT 500
            """, params
        ).fetchall()
        products_list = conn.execute("SELECT id, name FROM products ORDER BY name").fetchall()
    return render_template("stock.html", units=rows, q=q, status=status, products=products_list)


@app.route("/sales", methods=["GET", "POST"])
@login_required
@roles_required("admin", "finance", "sales", "support")
def sales():
    if request.method == "POST":
        order_number = request.form.get("order_number", "").strip()
        invoice_number = request.form.get("invoice_number", "").strip()
        channel = request.form.get("channel", "varejo")
        customer = request.form.get("customer", "").strip()
        sold_at = request.form.get("sold_at") or date.today().isoformat()
        notes = request.form.get("notes", "").strip()
        raw_chassis = request.form.get("chassis", "")
        chassis_list = [x.strip().upper() for x in raw_chassis.replace(";", ",").split(",") if x.strip()]

        if not order_number or not invoice_number:
            flash("Pedido e Nota Fiscal são campos obrigatórios.", "danger")
            return redirect(url_for("sales"))

        # Regra Negocial Inegociável: Chassi Obrigatório em todas as NFs de venda
        if not chassis_list and notes:
            # Tentar identificar se o chassi foi digitado na descrição/observação da NF
            with db() as conn:
                db_chassis = [r["chassis"].upper() for r in conn.execute("SELECT chassis FROM stock_units").fetchall() if r["chassis"]]
            found = [c for c in db_chassis if c in notes.upper()]
            if found:
                chassis_list = list(set(found))

        if not chassis_list:
            flash("Bloqueado: A Nota Fiscal (NF) deve conter obrigatoriamente um número de chassi válido na descrição e no cadastro da venda. Nenhuma nota é faturada sem chassi.", "danger")
            return redirect(url_for("sales"))

        if len(chassis_list) != len(set(chassis_list)):
            flash("Bloqueado: Há chassi duplicado na própria Nota Fiscal.", "danger")
            return redirect(url_for("sales"))

        # Garantir que a descrição (notes) da NF contenha explicitamente os números de chassi
        chassis_str = ", ".join(chassis_list)
        if not notes:
            notes = f"Chassi(s) da NF: {chassis_str}"
        elif not any(c in notes.upper() for c in chassis_list):
            notes = f"{notes} | Chassi(s) da NF: {chassis_str}"

        with db() as conn:
            if conn.execute("SELECT 1 FROM sales WHERE invoice_number=?", (invoice_number,)).fetchone():
                flash("Essa Nota Fiscal já foi cadastrada no sistema.", "danger")
                return redirect(url_for("sales"))

            units = []
            errors = []
            for ch in chassis_list:
                u = conn.execute(
                    """SELECT st.*,p.name product_name,p.retail_price,p.wholesale_price,i.status import_status
                       FROM stock_units st JOIN products p ON p.id=st.product_id LEFT JOIN imports i ON i.id=st.import_id
                       WHERE UPPER(st.chassis)=?""", (ch,)
                ).fetchone()
                if not u:
                    errors.append(f"{ch}: chassi não encontrado no estoque. Importe o lote antes de faturar a Nota Fiscal.")
                elif u["status"] != "available":
                    if u["status"] == "unreleased":
                        errors.append(f"{ch}: este chassi pertence a uma importação em conferência (não liberada pela Diretoria).")
                    else:
                        errors.append(f"{ch}: este chassi já consta como vendido ou indisponível (status: {u['status']}).")
                elif u["import_status"] != "released":
                    errors.append(f"{ch}: importação deste chassi ainda não foi liberada pela Diretoria.")
                else:
                    units.append(u)

            if errors:
                for e in errors:
                    flash(f"Bloqueio de Nota Fiscal — {e}", "danger")
                return redirect(url_for("sales"))

            danfe_file_obj = request.files.get("danfe_file")
            danfe_captured = request.form.get("danfe_captured_image")
            danfe_filename = None
            if danfe_captured:
                danfe_filename = save_base64_upload(danfe_captured, "danfe")
            elif danfe_file_obj and danfe_file_obj.filename:
                try:
                    danfe_filename = save_upload(danfe_file_obj, "danfe")
                except ValueError as e:
                    flash(f"Arquivo de comprovante da DANFE inválido: {str(e)}", "danger")
                    return redirect(url_for("sales"))

            # Processamento do Termo de Ciência da Garantia (Obrigatório no Varejo)
            warranty_term_file_obj = request.files.get("warranty_term_file")
            warranty_term_captured = request.form.get("warranty_term_captured_image")
            warranty_term_filename = None
            if warranty_term_captured:
                warranty_term_filename = save_base64_upload(warranty_term_captured, "garantia")
            elif warranty_term_file_obj and warranty_term_file_obj.filename:
                try:
                    warranty_term_filename = save_upload(warranty_term_file_obj, "garantia")
                except ValueError as e:
                    flash(f"Termo de Garantia inválido: {str(e)}", "danger")
                    return redirect(url_for("sales"))

            # Processamento do Canhoto da NF Assinado (Obrigatório no Atacado)
            signed_stub_file_obj = request.files.get("signed_stub_file")
            signed_stub_captured = request.form.get("signed_stub_captured_image")
            signed_stub_filename = None
            if signed_stub_captured:
                signed_stub_filename = save_base64_upload(signed_stub_captured, "canhoto")
            elif signed_stub_file_obj and signed_stub_file_obj.filename:
                try:
                    signed_stub_filename = save_upload(signed_stub_file_obj, "canhoto")
                except ValueError as e:
                    flash(f"Canhoto da NF inválido: {str(e)}", "danger")
                    return redirect(url_for("sales"))

            # Regra estrita de documentos por canal:
            if channel == "varejo" and not warranty_term_filename:
                flash("Para vendas no Varejo, é obrigatório anexar o Termo de Ciência da Garantia.", "danger")
                return redirect(url_for("sales"))
            elif channel == "atacado" and not signed_stub_filename:
                flash("Para vendas no Atacado, é obrigatório anexar o Canhoto da NF Assinado.", "danger")
                return redirect(url_for("sales"))

            # Processamento do Termo de Entrega (Permite múltiplas fotos / arquivos)
            term_filenames = []
            term_file_objs = request.files.getlist("delivery_term_files")
            for tf in term_file_objs:
                if tf and tf.filename:
                    try:
                        saved_name = save_upload(tf, "termo")
                        if saved_name:
                            term_filenames.append(saved_name)
                    except ValueError:
                        pass

            term_captured_raw = request.form.get("delivery_term_captured_images", "")
            if term_captured_raw:
                try:
                    cap_list = json.loads(term_captured_raw) if term_captured_raw.startswith("[") else [term_captured_raw]
                    for cap_b64 in cap_list:
                        if cap_b64 and "," in cap_b64:
                            saved_name = save_base64_upload(cap_b64, "termo")
                            if saved_name:
                                term_filenames.append(saved_name)
                except Exception:
                    pass

            # Processamento da Foto do Chassi (Plaqueta do Veículo / Caixa)
            chassis_photo_file_obj = request.files.get("chassis_photo_file")
            chassis_photo_captured = request.form.get("chassis_photo_captured_image")
            chassis_photo_filename = None
            if chassis_photo_captured:
                chassis_photo_filename = save_base64_upload(chassis_photo_captured, "chassis_veiculo")
            elif chassis_photo_file_obj and chassis_photo_file_obj.filename:
                try:
                    chassis_photo_filename = save_upload(chassis_photo_file_obj, "chassis_veiculo")
                except ValueError as e:
                    flash(f"Foto de chassi do veículo/caixa inválida: {str(e)}", "danger")
                    return redirect(url_for("sales"))

            if not danfe_filename and not term_filenames and not chassis_photo_filename:
                flash("É obrigatório anexar a Foto do Chassi do Veículo/Caixa, a DANFE ou o Termo de Entrega.", "danger")
                return redirect(url_for("sales"))

            term_files_json = json.dumps(term_filenames) if term_filenames else None
            vehicle_model = request.form.get("vehicle_model", "").strip()

            ai_verified = request.form.get("ai_chassis_verified") == "1"
            ai_extracted = request.form.get("ai_extracted_chassis", "").strip()

            # Bloqueio estrito no backend: se não veio validado pelo modal, valida agora nos documentos disponíveis
            if not ai_verified:
                target_filename = chassis_photo_filename or danfe_filename or warranty_term_filename or signed_stub_filename or (term_filenames[0] if term_filenames else None)
                if target_filename:
                    doc_path = UPLOAD_DIR / target_filename
                    if doc_path.exists():
                        ext = target_filename.rsplit(".", 1)[-1].lower()
                        mime = "application/pdf" if ext == "pdf" else f"image/{ext if ext != 'jpg' else 'jpeg'}"
                        v_res = extract_and_match_chassis(doc_path.read_bytes(), mime, chassis_list, expected_model=vehicle_model)
                        if not v_res.get("is_valid"):
                            flash(f"Bloqueio da IA: {v_res.get('summary', 'O chassi do documento não confere com o digitado.')}", "danger")
                            return redirect(url_for("sales"))
                        ai_verified = True
                        ai_extracted = ", ".join(v_res.get("extracted_chassis", []))

            default_total = sum(float(u["wholesale_price"] if channel == "atacado" else u["retail_price"]) for u in units)
            total_value = float(request.form.get("total_value") or default_total)
            cur = conn.execute(
                """INSERT INTO sales(order_number,invoice_number,channel,customer,sold_at,total_value,notes,danfe_file,delivery_term_files,vehicle_model,chassis_photo_file,warranty_term_file,signed_stub_file,ai_chassis_verified,ai_extracted_chassis,created_by)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (order_number, invoice_number, channel, customer, sold_at, total_value, notes, danfe_filename, term_files_json, vehicle_model, chassis_photo_filename, warranty_term_filename, signed_stub_filename, ai_verified, ai_extracted, session["user_id"]),
            )
            sale_id = cur.lastrowid
            per_unit = total_value / len(units) if units else 0
            for u in units:
                conn.execute("INSERT INTO sale_units(sale_id,stock_unit_id,product_id,unit_value) VALUES(?,?,?,?)", (sale_id, u["id"], u["product_id"], per_unit))
                conn.execute("UPDATE stock_units SET status='sold',sold_at=?,sale_id=? WHERE id=?", (sold_at, sale_id, u["id"]))

            methods = request.form.getlist("payment_method[]")
            accounts = request.form.getlist("payment_account[]")
            amounts = request.form.getlist("payment_amount[]")
            receipt_files = request.files.getlist("payment_receipt[]")

            sale_photo_receipt = save_base64_upload(request.form.get("captured_image_data"), "venda")
            if not sale_photo_receipt:
                try:
                    sale_photo_receipt = save_upload(request.files.get("sale_receipt_file"), "venda")
                except ValueError:
                    sale_photo_receipt = None

            receipt_total = 0.0
            for idx, amount_str in enumerate(amounts):
                if not amount_str:
                    continue
                amount = float(amount_str)
                receipt_total += amount
                receipt = None
                if idx < len(receipt_files) and receipt_files[idx] and receipt_files[idx].filename:
                    try:
                        receipt = save_upload(receipt_files[idx], "recebimento")
                    except ValueError:
                        receipt = None
                if not receipt and idx == 0 and sale_photo_receipt:
                    receipt = sale_photo_receipt

                method = methods[idx] if idx < len(methods) else "À vista"
                account = accounts[idx] if idx < len(accounts) else ""
                conn.execute("INSERT INTO sale_receipts(sale_id,method,account,amount,received_at,receipt_file) VALUES(?,?,?,?,?,?)", (sale_id, method, account, amount, sold_at, receipt))

            if not amounts and sale_photo_receipt:
                conn.execute("INSERT INTO sale_receipts(sale_id,method,account,amount,received_at,receipt_file) VALUES(?,?,?,?,?,?)", (sale_id, "À vista", "Geral", total_value, sold_at, sale_photo_receipt))

            conn.commit()
        audit("sale.created", f"sale_id={sale_id}; invoice={invoice_number}; chassis={','.join(chassis_list)}")
        if abs(receipt_total - total_value) > 0.01:
            flash(f"Venda registrada, mas os recebimentos somam {money(receipt_total)} e a venda {money(total_value)}. Confira.", "warning")
        else:
            flash("Venda registrada e chassis baixados sem duplicidade.", "success")
        return redirect(url_for("sales"))

    with db() as conn:
        sales_rows = conn.execute(
            """
            SELECT s.*, u.name created_by_name, COUNT(DISTINCT su.id) units, COALESCE(SUM(sr.amount),0) received
            FROM sales s 
            LEFT JOIN users u ON u.id=s.created_by
            LEFT JOIN sale_units su ON su.sale_id=s.id
            LEFT JOIN sale_receipts sr ON sr.sale_id=s.id
            GROUP BY s.id, u.name ORDER BY s.sold_at DESC, s.id DESC LIMIT 200
            """
        ).fetchall()
        sales_data = []
        if sales_rows:
            sale_ids = [s['id'] for s in sales_rows]
            ids_str = ','.join(str(x) for x in sale_ids)
            chassis_rows = conn.execute(
                f"""SELECT su.sale_id, st.chassis, p.name product_name FROM sale_units su
                   JOIN stock_units st ON st.id=su.stock_unit_id
                   JOIN products p ON p.id=su.product_id
                   WHERE su.sale_id IN ({ids_str})"""
            ).fetchall()
            receipts_rows = conn.execute(
                f"""SELECT * FROM sale_receipts WHERE sale_id IN ({ids_str})"""
            ).fetchall()
            
            chassis_by_sale = {}
            for c in chassis_rows:
                chassis_by_sale.setdefault(c['sale_id'], []).append(dict(c))
                
            receipts_by_sale = {}
            for r in receipts_rows:
                receipts_by_sale.setdefault(r['sale_id'], []).append(dict(r))
                
            for s in sales_rows:
                sd = dict(s)
                c_list = chassis_by_sale.get(sd['id'], [])
                r_list = receipts_by_sale.get(sd['id'], [])
                sd['chassis_details'] = c_list
                sd['chassis_str'] = ', '.join(c['chassis'] for c in c_list)
                sd['receipts'] = r_list
                t_files = []
                if sd.get('delivery_term_files'):
                    try:
                        raw_t = sd['delivery_term_files']
                        t_files = json.loads(raw_t) if isinstance(raw_t, str) and raw_t.startswith("[") else ([raw_t] if raw_t else [])
                    except Exception:
                        t_files = []
                sd['term_files_list'] = t_files
                sales_data.append(sd)
    return render_template("sales.html", sales=sales_data)


@app.route("/payments", methods=["GET", "POST"])
@login_required
@roles_required("admin", "finance", "support")
def payments():
    if request.method == "POST":
        paid_at = request.form.get("paid_at") or date.today().isoformat()
        description = request.form.get("description", "").strip()
        amount = float(request.form.get("amount") or 0)
        category = request.form.get("category", "").strip()
        account = request.form.get("account", "").strip()
        payment_method = request.form.get("payment_method", "").strip()
        card_last4 = request.form.get("card_last4", "").strip()
        supplier = request.form.get("supplier", "").strip()
        document_no = request.form.get("document_no", "").strip()
        import_id = request.form.get("import_id") or None
        visibility = "admin_only" if import_id else "finance"

        try:
            receipt = save_base64_upload(request.form.get("captured_image_data"), "pagamento")
            if not receipt:
                receipt = save_upload(request.files.get("receipt_file"), "pagamento")
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("payments"))

        # Exigência estrita de comprovante/recibo
        if not receipt:
            flash("⚠️ É OBRIGATÓRIO anexar a Nota Fiscal do Fornecedor ou Recibo de Pagamento (por foto ou arquivo).", "danger")
            return redirect(url_for("payments"))

        if not description or amount <= 0:
            flash("Descrição e valor são obrigatórios.", "danger")
            return redirect(url_for("payments"))

        with db() as conn:
            conn.execute(
                """INSERT INTO payments(paid_at, description, category, amount, account, payment_method, card_last4, supplier, document_no, receipt_file, import_id, visibility, created_by)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (paid_at, description, category, amount, account, payment_method, card_last4, supplier, document_no, receipt, import_id, visibility, session["user_id"]),
            )
            conn.commit()
        audit("payment.created", f"{description}; R$ {amount}; Categoria: {category}; Conta: {account}")
        flash("🎉 Pagamento realizado registrado com sucesso!", "success")
        return redirect(url_for("payments"))

    u = current_user()
    with db() as conn:
        if u["role"] in ("admin", "support"):
            rows = conn.execute("SELECT p.*,i.reference import_ref FROM payments p LEFT JOIN imports i ON i.id=p.import_id ORDER BY p.paid_at DESC,p.id DESC LIMIT 300").fetchall()
            imports_rows = conn.execute("SELECT id,reference FROM imports ORDER BY created_at DESC").fetchall()
        else:
            rows = conn.execute("SELECT p.*,NULL import_ref FROM payments p WHERE visibility='finance' ORDER BY p.paid_at DESC,p.id DESC LIMIT 300").fetchall()
            imports_rows = []
    return render_template(
        "payments.html",
        payments=rows,
        imports=imports_rows,
        categories=PAYMENT_CATEGORIES,
        accounts=ACCOUNTS_LIST,
        payment_methods=PAYMENT_METHODS
    )


@app.route("/api/payments/analyze-receipt", methods=["POST"])
@login_required
def api_analyze_payment_receipt():
    try:
        image_bytes = None
        mime_type = "image/jpeg"

        base64_str = request.form.get("captured_image_data")
        if base64_str and "," in base64_str:
            header, data = base64_str.split(",", 1)
            image_bytes = base64.b64decode(data)
            mime_type = "image/png" if "png" in header else "image/jpeg"
        elif "receipt_file" in request.files:
            f = request.files["receipt_file"]
            image_bytes = f.read()
            mime_type = f.content_type or "image/jpeg"

        if not image_bytes:
            return jsonify({"success": False, "message": "Nenhum arquivo ou foto foi enviado para análise."}), 400

        form_data = {
            "amount": request.form.get("amount"),
            "paid_at": request.form.get("paid_at"),
            "category": request.form.get("category"),
            "account": request.form.get("account"),
            "payment_method": request.form.get("payment_method"),
            "card_last4": request.form.get("card_last4")
        }

        analysis = analyze_payment_receipt(image_bytes, mime_type=mime_type, form_data=form_data)
        return jsonify(analysis)
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 400


@app.route("/users", methods=["GET", "POST"])
@login_required
@roles_required("admin", "support")
def users():
    if request.method == "POST":
        name = request.form["name"].strip()
        username = request.form["username"].strip().lower()
        password = request.form.get("password") or "MOne2026!"
        role = request.form.get("role", "sales")
        with db() as conn:
            try:
                conn.execute("INSERT INTO users(name,username,password_hash,role) VALUES(?,?,?,?)", (name, username, hash_password(password), role))
                conn.commit()
                flash("Usuário criado.", "success")
            except Exception:
                flash("Esse nome de usuário já existe.", "danger")
        return redirect(url_for("users"))
    with db() as conn:
        rows = conn.execute("SELECT * FROM users ORDER BY name").fetchall()
    return render_template("users.html", users=rows)

@app.route("/users/<int:uid>/toggle", methods=["POST"])
@login_required
@roles_required("admin", "support")
def toggle_user(uid):
    with db() as conn:
        u = conn.execute("SELECT active FROM users WHERE id=?", (uid,)).fetchone()
        if u:
            new_val = 0 if u["active"] else 1
            conn.execute("UPDATE users SET active=? WHERE id=?", (new_val, uid))
            conn.commit()
            flash("Status do usuário alterado.", "success")
    return redirect(url_for("users"))


@app.route("/users/<int:uid>/reset-password", methods=["POST"])
@login_required
@roles_required("admin", "support")
def reset_user_password(uid):
    with db() as conn:
        conn.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password("MOne2026!"), uid))
        conn.commit()
        flash("Senha resetada para MOne2026!.", "success")
    return redirect(url_for("users"))


@app.route("/users/<int:uid>/edit", methods=["POST"])
@login_required
@roles_required("admin", "support")
def edit_user(uid):
    name = request.form.get("name", "").strip()
    username = request.form.get("username", "").strip().lower()
    role = request.form.get("role", "sales")
    new_password = request.form.get("password", "").strip()
    with db() as conn:
        if new_password:
            conn.execute(
                "UPDATE users SET name=?, username=?, role=?, password_hash=? WHERE id=?",
                (name, username, role, hash_password(new_password), uid)
            )
        else:
            conn.execute(
                "UPDATE users SET name=?, username=?, role=? WHERE id=?",
                (name, username, role, uid)
            )
        conn.commit()
    audit("user.edited", f"user_id={uid}; username={username}; role={role}")
    flash("Usuário atualizado com sucesso.", "success")
    return redirect(url_for("users"))


@app.route("/users/<int:uid>/delete", methods=["POST"])
@login_required
@roles_required("admin")
def delete_user(uid):
    if uid == session.get("user_id"):
        flash("Você não pode excluir a sua própria conta logada.", "danger")
        return redirect(url_for("users"))
    try:
        with db() as conn:
            # Desvincula o usuário dos registros vinculados mantendo o histórico de vendas/importações intacto
            conn.execute("UPDATE imports SET created_by=NULL WHERE created_by=?", (uid,))
            conn.execute("UPDATE sales SET created_by=NULL WHERE created_by=?", (uid,))
            conn.execute("UPDATE payments SET created_by=NULL WHERE created_by=?", (uid,))
            conn.execute("UPDATE audit_log SET user_id=NULL WHERE user_id=?", (uid,))
            conn.execute("DELETE FROM users WHERE id=?", (uid,))
            conn.commit()
        audit("user.deleted", f"user_id={uid}")
        flash("Usuário excluído com sucesso.", "success")
    except Exception as e:
        flash(f"Erro ao excluir usuário: {str(e)}", "danger")
    return redirect(url_for("users"))




@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    new_password = request.form.get("new_password", "")
    if len(new_password) < 8:
        flash("Use uma senha com pelo menos 8 caracteres.", "danger")
        return redirect(request.referrer or url_for("dashboard"))
    with db() as conn:
        conn.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(new_password), session["user_id"]))
        conn.commit()
    flash("Senha alterada.", "success")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/uploads/<path:filename>")
@login_required
def uploads(filename):
    return send_from_directory(UPLOAD_DIR, filename, as_attachment=False)


@app.route("/api/chassis/<path:chassis>")
@login_required
def api_chassis(chassis):
    with db() as conn:
        row = conn.execute(
            """SELECT st.chassis,st.status,st.color,st.motor_no,p.name product,i.reference import_ref,i.status import_status,s.invoice_number
               FROM stock_units st JOIN products p ON p.id=st.product_id LEFT JOIN imports i ON i.id=st.import_id LEFT JOIN sales s ON s.id=st.sale_id
               WHERE st.chassis=?""", (chassis,)
        ).fetchone()
    if not row:
        return jsonify({"ok": False, "message": "Chassi não encontrado. Importe a planilha do contêiner."}), 404
    return jsonify({"ok": True, "unit": dict(row)})




@app.route("/stock/add", methods=["POST"])
@login_required
@roles_required("admin", "stock", "support")
def add_stock_unit():
    chassis = request.form.get("chassis", "").strip()
    product_id = request.form.get("product_id")
    color = request.form.get("color", "").strip()
    motor_no = request.form.get("motor_no", "").strip()
    location = request.form.get("location", "Depósito").strip()
    received_at = request.form.get("received_at") or date.today().isoformat()
    if not chassis or not product_id:
        flash("Chassi e produto são obrigatórios.", "danger")
        return redirect(url_for("stock"))
    with db() as conn:
        if conn.execute("SELECT 1 FROM stock_units WHERE chassis=?", (chassis,)).fetchone():
            flash("Chassi já cadastrado na base.", "danger")
            return redirect(url_for("stock"))
        conn.execute(
            """INSERT INTO stock_units(chassis,motor_no,product_id,color,status,location,received_at)
               VALUES(?,?,?,?,'available',?,?)""",
            (chassis, motor_no, product_id, color, location, received_at)
        )
        conn.commit()
    audit("stock.unit_added", f"chassis={chassis}")
    flash("Unidade adicionada ao estoque com sucesso.", "success")
    return redirect(url_for("stock"))


@app.route("/stock/export")
@login_required
def export_stock():
    with db() as conn:
        rows = conn.execute(
            """SELECT st.chassis, p.name product_name, st.color, st.motor_no, COALESCE(i.reference, 'Nacional') import_ref, st.location, st.status, st.received_at
               FROM stock_units st JOIN products p ON p.id=st.product_id LEFT JOIN imports i ON i.id=st.import_id
               ORDER BY st.created_at DESC"""
        ).fetchall()
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["CHASSI", "PRODUTO", "COR", "MOTOR", "IMPORTACAO", "LOCAL", "STATUS", "DATA_ENTRADA"])
    for r in rows:
        writer.writerow([r["chassis"], r["product_name"], r["color"] or "", r["motor_no"] or "", r["import_ref"], r["location"], r["status"], r["received_at"] or ""])
    return out.getvalue(), 200, {"Content-Type": "text/csv; charset=utf-8-sig", "Content-Disposition": "attachment; filename=estoque_m_one.csv"}


@app.route("/sales/export")
@login_required
def export_sales():
    with db() as conn:
        rows = conn.execute(
            """SELECT s.sold_at, s.order_number, s.invoice_number, s.channel, s.customer, s.total_value,
                      (SELECT GROUP_CONCAT(st.chassis, ', ') FROM sale_units su JOIN stock_units st ON st.id=su.stock_unit_id WHERE su.sale_id=s.id) chassis_list
               FROM sales s ORDER BY s.sold_at DESC"""
        ).fetchall()
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["DATA", "PEDIDO_BLING", "NOTA_FISCAL", "CANAL", "CLIENTE", "VALOR_TOTAL", "CHASSIS"])
    for r in rows:
        writer.writerow([r["sold_at"], r["order_number"], r["invoice_number"], r["channel"], r["customer"] or "", f"{r['total_value']:.2f}".replace(".", ","), r["chassis_list"] or ""])
    return out.getvalue(), 200, {"Content-Type": "text/csv; charset=utf-8-sig", "Content-Disposition": "attachment; filename=vendas_m_one.csv"}


@app.route("/payments/export")
@login_required
@roles_required("admin", "finance", "support")
def export_payments():
    u = current_user()
    with db() as conn:
        if u["role"] in ("admin", "support"):
            rows = conn.execute(
                """SELECT p.paid_at, p.description, p.category, p.account, p.amount, i.reference import_ref
                   FROM payments p LEFT JOIN imports i ON i.id=p.import_id ORDER BY p.paid_at DESC"""
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT p.paid_at, p.description, p.category, p.account, p.amount, NULL import_ref
                   FROM payments p WHERE visibility='finance' ORDER BY p.paid_at DESC"""
            ).fetchall()
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["DATA", "DESCRICAO", "CATEGORIA", "CONTA", "VALOR", "IMPORTACAO"])
    for r in rows:
        writer.writerow([r["paid_at"], r["description"], r["category"] or "", r["account"] or "", f"{r['amount']:.2f}".replace(".", ","), r["import_ref"] or ""])
    return out.getvalue(), 200, {"Content-Type": "text/csv; charset=utf-8-sig", "Content-Disposition": "attachment; filename=pagamentos_m_one.csv"}

@app.errorhandler(413)
def too_large(_):
    flash("Arquivo grande demais. Limite: 20 MB.", "danger")
    return redirect(request.referrer or url_for("dashboard"))

@app.errorhandler(500)
def handle_500(e):
    import traceback
    return f"<h3>Erro Interno</h3><pre>{traceback.format_exc()}</pre>", 500


@app.route("/copilot")
@login_required
def copilot():
    has_key = bool(get_gemini_api_key())
    return render_template("copilot.html", has_key=has_key)


@app.route("/api/copilot/chat", methods=["POST"])
@login_required
def copilot_chat():
    data = request.get_json(silent=True) or {}
    message = (data.get("message") or "").strip()
    history = data.get("history") or []
    if not message:
        return jsonify({"success": False, "message": "Mensagem não informada."}), 400

    u = current_user()
    if not u:
        return jsonify({"success": False, "message": "Sessão expirada. Faça login novamente."}), 401

    with db() as conn:
        result = ask_gemini_copilot(
            user_message=message,
            history=history,
            db_conn=conn,
            user_role=u["role"],
            user_name=u["name"],
        )
    return jsonify(result)


@app.route("/api/sales/verify-chassis", methods=["POST"])
@login_required
def verify_sales_chassis():
    chassis_raw = request.form.get("chassis", "")
    chassis_list = [x.strip() for x in chassis_raw.replace(";", ",").split(",") if x.strip()]
    if not chassis_list:
        return jsonify({
            "success": False,
            "is_valid": False,
            "message": "Nenhum número de chassi informado. Digite o(s) chassi(s) no formulário antes de conferir o comprovante."
        }), 400

    file_objs = request.files.getlist("file")
    captured_data = request.form.get("captured_image")
    captured_list_raw = request.form.get("captured_images")

    items_to_check = []

    if captured_data and "," in captured_data:
        try:
            header, data_str = captured_data.split(",", 1)
            b = base64.b64decode(data_str)
            m = "image/png" if "png" in header else ("image/webp" if "webp" in header else "image/jpeg")
            items_to_check.append((b, m))
        except Exception:
            pass

    if captured_list_raw:
        try:
            cap_arr = json.loads(captured_list_raw) if captured_list_raw.startswith("[") else [captured_list_raw]
            for c_str in cap_arr:
                if c_str and "," in c_str:
                    header, data_str = c_str.split(",", 1)
                    b = base64.b64decode(data_str)
                    m = "image/png" if "png" in header else ("image/webp" if "webp" in header else "image/jpeg")
                    items_to_check.append((b, m))
        except Exception:
            pass

    chassis_photo_file = request.files.get("chassis_photo_file")
    chassis_photo_captured = request.form.get("chassis_photo_captured_image")

    if chassis_photo_captured and "," in chassis_photo_captured:
        try:
            header, data_str = chassis_photo_captured.split(",", 1)
            b = base64.b64decode(data_str)
            m = "image/png" if "png" in header else ("image/webp" if "webp" in header else "image/jpeg")
            items_to_check.append((b, m))
        except Exception:
            pass

    if chassis_photo_file and chassis_photo_file.filename:
        ext = chassis_photo_file.filename.rsplit(".", 1)[-1].lower()
        if ext in ALLOWED_EXTENSIONS:
            b = chassis_photo_file.read()
            m = "application/pdf" if ext == "pdf" else ("image/png" if ext == "png" else "image/jpeg")
            items_to_check.append((b, m))

    for key_file, key_cap in [("warranty_term_file", "warranty_term_captured_image"), ("signed_stub_file", "signed_stub_captured_image")]:
        cap_val = request.form.get(key_cap)
        if cap_val and "," in cap_val:
            try:
                header, data_str = cap_val.split(",", 1)
                b = base64.b64decode(data_str)
                m = "image/png" if "png" in header else ("image/webp" if "webp" in header else "image/jpeg")
                items_to_check.append((b, m))
            except Exception:
                pass
        f_val = request.files.get(key_file)
        if f_val and f_val.filename:
            ext = f_val.filename.rsplit(".", 1)[-1].lower()
            if ext in ALLOWED_EXTENSIONS:
                b = f_val.read()
                m = "application/pdf" if ext == "pdf" else ("image/png" if ext == "png" else "image/jpeg")
                items_to_check.append((b, m))

    if not items_to_check:
        return jsonify({"success": False, "is_valid": False, "message": "Nenhum arquivo ou foto foi anexado para a conferência."}), 400

    vehicle_model = request.form.get("vehicle_model", "").strip()

    # Passa todos os comprovantes anexados para a IA realizar a triangulação de chassi
    res = extract_and_match_chassis(items_to_check, chassis_list, expected_model=vehicle_model)
    return jsonify(res)


@app.route("/integrations/bling", methods=["GET"])
@login_required
@roles_required("admin", "support")
def integrations_bling():
    rec = bling_service.get_bling_integration_record()
    has_credentials = bool(rec and rec.get("client_id") and rec.get("client_secret"))
    is_connected = bool(rec and rec.get("access_token"))
    callback_url = url_for("bling_callback", _external=True)
    return render_template(
        "integrations_bling.html",
        rec=rec or {},
        has_credentials=has_credentials,
        is_connected=is_connected,
        callback_url=callback_url
    )


@app.route("/integrations/bling/save", methods=["POST"])
@login_required
@roles_required("admin", "support")
def save_bling_config():
    client_id = request.form.get("client_id", "").strip()
    client_secret = request.form.get("client_secret", "").strip()
    if not client_id or not client_secret:
        flash("Informe o Client ID e o Client Secret do Bling.", "danger")
        return redirect(url_for("integrations_bling"))
    bling_service.save_bling_credentials(client_id, client_secret)
    flash("Credenciais do Bling salvas com sucesso! Agora clique em 'Conectar com Bling'.", "success")
    return redirect(url_for("integrations_bling"))


@app.route("/bling/authorize")
@login_required
@roles_required("admin", "support")
def bling_authorize():
    try:
        callback_url = url_for("bling_callback", _external=True)
        auth_url = bling_service.get_bling_auth_url(callback_url)
        return redirect(auth_url)
    except Exception as e:
        flash(f"Erro ao iniciar autorização com Bling: {str(e)}", "danger")
        return redirect(url_for("integrations_bling"))


@app.route("/bling/callback")
@login_required
def bling_callback():
    code = request.args.get("code")
    err = request.args.get("error")
    if err:
        flash(f"Autorização cancelada ou recusada no Bling: {err}", "danger")
        return redirect(url_for("integrations_bling"))
    if not code:
        flash("Nenhum código de autorização retornado pelo Bling.", "danger")
        return redirect(url_for("integrations_bling"))
    try:
        callback_url = url_for("bling_callback", _external=True)
        bling_service.exchange_code_for_token(code, callback_url)
        flash("🎉 Conexão com o Bling ERP autorizada e ativada com sucesso!", "success")
    except Exception as e:
        flash(f"Falha ao trocar código pelo token do Bling: {str(e)}", "danger")
    return redirect(url_for("integrations_bling"))


@app.route("/bling/disconnect")
@login_required
@roles_required("admin", "support")
def bling_disconnect():
    with db() as conn:
        conn.execute("UPDATE integrations SET access_token = NULL, refresh_token = NULL WHERE service_name = 'bling'")
        conn.commit()
    flash("Conexão com o Bling foi desconectada.", "warning")
    return redirect(url_for("integrations_bling"))


@app.route("/api/bling/order/<path:order_num>")
@login_required
def api_bling_order(order_num):
    try:
        data = bling_service.search_bling_order(order_num)
        return jsonify(data)
    except Exception as e:
        return jsonify({"found": False, "message": str(e)}), 400


@app.route("/api/bling/invoice/<path:inv_num>")
@login_required
def api_bling_invoice(inv_num):
    try:
        data = bling_service.search_bling_invoice(inv_num)
        return jsonify(data)
    except Exception as e:
        return jsonify({"found": False, "message": str(e)}), 400


@app.route("/products/sync-bling-stock", methods=["POST"])
@app.route("/api/bling/sync-stock", methods=["POST"])
@login_required
def sync_bling_stock_route():
    try:
        res = bling_service.sync_bling_products_stock()
        if request.headers.get("Accept") == "application/json" or request.is_json:
            return jsonify(res)
        if res.get("success"):
            flash(f"✅ {res.get('message')}", "success")
        else:
            flash(f"⚠️ {res.get('message')}", "warning")
    except Exception as e:
        if request.headers.get("Accept") == "application/json" or request.is_json:
            return jsonify({"success": False, "message": str(e)}), 400
        flash(f"Erro ao sincronizar estoque com o Bling: {str(e)}", "danger")
    return redirect(url_for("products"))


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5001")), debug=True)


